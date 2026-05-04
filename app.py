import os
import re
import pandas as pd
from flask import Flask, request, send_file, render_template
from openpyxl.styles import Font

app = Flask(__name__)

UPLOAD_FOLDER = 'uploads'
os.makedirs(UPLOAD_FOLDER, exist_ok=True)


# -------------------------------
# PROCESS DATA
# -------------------------------
def process_advancements(file_path):
    ext = file_path.split('.')[-1].lower()

    if ext == 'xlsx':
        df = pd.read_excel(file_path, engine='openpyxl', header=4)
    elif ext == 'xls':
        df = pd.read_excel(file_path, engine='xlrd', header=4)
    elif ext == 'csv':
        df = pd.read_csv(file_path)
    else:
        raise ValueError("Unsupported file format")

    # Normalize columns
    df.columns = df.columns.str.strip().str.upper()

    column_mapping = {
        'DESCRIPTION': 'Description',
        'DATE': 'Date',
        'DEBIT': 'Amount'
    }

    df = df.rename(columns=column_mapping)

    required = ['Description', 'Amount', 'Date']
    for col in required:
        if col not in df.columns:
            raise ValueError(f"Missing column: {col}")

    # Clean data
    df['Date'] = pd.to_datetime(df['Date'], dayfirst=True, errors='coerce')
    df['Amount'] = pd.to_numeric(df['Amount'], errors='coerce').fillna(0)

    # Keep only advancement rows
    df = df[df['Description'].str.contains(r"(?i)salary\s+advancement", na=False)]


    # -------------------------------
    # NAME NORMALIZATION
    # -------------------------------
    def extract_name(desc):
        if pd.isna(desc):
            return "UNKNOWN"

        text = str(desc).strip().lower()

        # Fix common typo
        text = re.sub(r"\bfpr\b", "for", text)

        # Normalize spaces
        text = re.sub(r"\s+", " ", text)

        # Remove prefix
        name = re.sub(r"salary\s+advancement\s+for\s+", "", text).strip()

        # Remove punctuation (handles EVERLINE')
        name = re.sub(r"[^\w\s]", "", name)

        # First name only
        first_name = name.split()[0] if name else "unknown"
        first_name = first_name.upper()

        # -------------------------------
        # MERGE NAME VARIANTS
        # -------------------------------
        alias_map = {
            # TERVIL variants
            "TERVIIL": "TERVIL",
            "TERVILL": "TERVIL",

            # EVERLINE variants
            "EVERLYN": "EVERLINE",
            "EVERLINE": "EVERLINE",
            "EVERLINEE": "EVERLINE",
            "EVERLIINE": "EVERLINE",
            "EVERLIN": "EVERLINE"
        }

        return alias_map.get(first_name, first_name)

    df['Employee Name'] = df['Description'].apply(extract_name)

    grouped = df.groupby('Employee Name')

    results = []

    for name, group in grouped:
        sorted_group = group.sort_values('Date')

        results.append({
            "employee_name": name,
            "total": sorted_group['Amount'].sum(),
            "advancements": sorted_group[['Date', 'Amount']].to_dict('records')
        })

    return results


# -------------------------------
# EXPORT REPORT
# -------------------------------
def export_report(results, output_path):
    report_rows = []
    grand_total = 0

    bold_font = Font(bold=True)

    for emp in results:

        # Employee header
        report_rows.append({
            "Date": f"--- EMPLOYEE: {emp['employee_name']} ---",
            "Description": "",
            "Amount": "",
            "Running Total": ""
        })

        running_total = 0

        for adv in emp['advancements']:
            running_total += adv['Amount']
            clean_date = adv['Date'].strftime('%Y-%m-%d') if pd.notnull(adv['Date']) else "N/A"

            report_rows.append({
                "Date": clean_date,
                "Description": f"Advancement for {emp['employee_name']}",
                "Amount": float(adv['Amount']),
                "Running Total": float(running_total)
            })

        # TOTAL row
        report_rows.append({
            "Date": "",
            "Description": f"TOTAL FOR {emp['employee_name']}",
            "Amount": "",
            "Running Total": float(emp['total'])
        })

        # spacer
        report_rows.append({
            "Date": "",
            "Description": "",
            "Amount": "",
            "Running Total": ""
        })

        grand_total += emp['total']

    # GRAND TOTAL
    report_rows.append({
        "Date": "GRAND TOTAL",
        "Description": "Total of all advancements",
        "Amount": "",
        "Running Total": float(grand_total)
    })

    df_output = pd.DataFrame(report_rows)

    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        df_output.to_excel(writer, index=False, sheet_name='Salary Advancements')

        worksheet = writer.sheets['Salary Advancements']

        # Column widths
        for idx, col in enumerate(df_output.columns):
            max_len = df_output[col].astype(str).map(len).max()
            header_len = len(str(col))
            if idx < 26:
                worksheet.column_dimensions[chr(65 + idx)].width = max(max_len, header_len) + 5

        # Bold formatting
        for row_idx, row in enumerate(report_rows, start=2):

            # Employee header
            if isinstance(row["Date"], str) and row["Date"].startswith("--- EMPLOYEE"):
                for col in range(1, 5):
                    worksheet.cell(row=row_idx, column=col).font = bold_font

            # TOTAL rows
            if isinstance(row["Description"], str) and "TOTAL FOR" in row["Description"]:
                for col in range(1, 5):
                    worksheet.cell(row=row_idx, column=col).font = bold_font

            # GRAND TOTAL
            if row["Date"] == "GRAND TOTAL":
                for col in range(1, 5):
                    worksheet.cell(row=row_idx, column=col).font = bold_font


# -------------------------------
# ROUTES
# -------------------------------
@app.route('/', methods=['GET', 'POST'])
def index():
    if request.method == 'POST':
        file = request.files.get('file')

        if not file or file.filename == '':
            return "No file uploaded"

        filepath = os.path.join(UPLOAD_FOLDER, file.filename)
        file.save(filepath)

        try:
            results = process_advancements(filepath)

            output_file = os.path.join(UPLOAD_FOLDER, 'Advancement_Report.xlsx')
            export_report(results, output_file)

            return send_file(output_file, as_attachment=True)

        except Exception as e:
            return f"Error: {str(e)}"

    return render_template('index.html')


# -------------------------------
# RUN APP
# -------------------------------
if __name__ == '__main__':
    app.run(debug=True)